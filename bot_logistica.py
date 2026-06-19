import os
import telebot
from openpyxl import load_workbook, Workbook

#Configuración inicial 

TOKEN = ""
bot = telebot.TeleBot(TOKEN)
DIRECTORIO_ACTUAL = os.path.dirname(os.path.abspath(__file__))
ARCHIVO_EXCEL = os.path.join(DIRECTORIO_ACTUAL, "base_datos_aduanas.xlsx")

#Base de datos

def actualizar_estado_excel (usuario_id, nuevo_estado, nombre_usuario="Usuario"):
    """Actualiza la Máquina de Estados en el Excel o crea un usuario nuevo."""
    if not os.path.exists (ARCHIVO_EXCEL):
        wb = Workbook()
        hoja = wb.active
        hoja.append(["usuario_id", "nombre_usuario", "estado_actual", "cupos_courier_usados", "monto_franquicia_usado"])
        wb.save(ARCHIVO_EXCEL)

    wb = load_workbook (ARCHIVO_EXCEL)
    hoja = wb.active
    usuario_encontrado = False

    for fila in range(2, hoja.max_row + 1):
        celda_id = hoja.cell(row=fila, column=1).value
        if str(celda_id) == str(usuario_id):
            hoja.cell(row=fila, column=3).value = nuevo_estado
            usuario_encontrado = True
            break

    if not usuario_encontrado:
        hoja.append([usuario_id, nombre_usuario, nuevo_estado, 0, 0.0])

    wb.save(ARCHIVO_EXCEL)

def leer_datos_usuario (usuario_id):
    """Lee el estado actual y los consumos aduaneros del usuario."""
    if not os.path.exists (ARCHIVO_EXCEL):
        return {"estado_actual": "inicio", "cupos_usados": 0, "monto_usado": 0.0}
    
    wb = load_workbook (ARCHIVO_EXCEL)
    hoja = wb.active

    for fila in range(2, hoja.max_row + 1):
        celda_id = hoja.cell(row=fila, column=1).value
        if str(celda_id) == str(usuario_id):
            return {"estado_actual": hoja.cell(row=fila, column=3).value, "cupos_usados": hoja.cell(row=fila, column=4).value, "monto_usado": float(hoja.cell(row=fila, column=5).value)}
    
    return {"estado_actual": "inicio", "cupos_usados": 0, "monto_usado": 0.0}

#Interacción con Bot y Compuertas BPMN

@bot.message_handler(commands=['start', 'hola'])

def enviar_bienvenida(message):
    usuario_id = message.chat.id
    nombre = message.chat.first_name

    #Estado inicial en base de datos
    actualizar_estado_excel (usuario_id, "esperando_opcion_menu", nombre)

    texto_bienvenida = (f"¡Hola, {nombre}! Soy tu asistente de logística aduanera. \n\n"
        "¿En qué te puedo ayudar hoy? Ingresá el número de la opción:\n"
        "1. Simular cálculo de impuestos de importación.\n"
        "2. Consultar cupos anuales disponibles.\n"
        "3. Diferencias entre Courier Privado y Correo Nacional.")
    bot.reply_to (message, texto_bienvenida)

@bot.message_handler(func=lambda message: True)

def procesar_mensajes(message):
    texto_usuario = message.text.strip()
    usuario_id = message.chat.id

    #Se lee el estado del usuario
    datos_usuario = leer_datos_usuario (usuario_id)
    estado = datos_usuario ["estado_actual"]

    #Derivación teniendo en cuenta el estado 
    if estado == "esperando_opcion_menu":
        if texto_usuario == "1":
            bot.reply_to(message, "Perfecto. Por favor ingrese el valor total de tu compra en dólares (ej: 450 o 450.50):")
            actualizar_estado_excel(usuario_id, 'esperando_monto')
        
        elif texto_usuario == "2":
            cupos = datos_usuario["cupos_usados"]
            bot.reply_to(message, f"Has utilizado {cupos} de tus 5 cupos anuales de Courier.")
            actualizar_estado_excel(usuario_id, 'inicio')

        elif texto_usuario == "3":
            info_correo = (
                "*Courier Privado* (DHL, FedEx, etc):\n"
                "Límite de $1000 USD y 50kg. La empresa hace el trámite de aduana por vos.\n\n"
                "*Correo Nacional*:\n"
                "Límite de $3000 USD y 20kg. Tenés que hacer la declaración vos mismo en el portal de envíos internacionales (Epago)."
            )
            bot.reply_to(message, info_correo, parse_mode="Markdown")
            bot.send_message(usuario_id, "Escribí /start para volver al menú principal.")
            actualizar_estado_excel(usuario_id, 'inicio')

        else:
            bot.reply_to(message, "Por favor, ingresá una opción válida (1, 2 o 3).")

    elif estado == "esperando_monto":
        try:
            monto_compra = float (texto_usuario)

            #Compuerta Exclusiva
            if monto_compra <= 1000:
                #Camino feliz (Sí)
                bot.reply_to(message, f"El monto de ${monto_compra} USD se encuentra dentro del límite para el régimen Courier.")
                
                cupos = datos_usuario["cupos_usados"]
                if cupos < 5:
                    bot.send_message(usuario_id, f"Tenés {5 - cupos} cupos disponibles. El trámite se encuentra pre-aprobado.")
                else:
                    bot.send_message(usuario_id, " Alcanzaste el límite de 5 vuelos anuales. No podés usar Courier.")
                
            else:
                #Camino infeliz (No)
                bot.reply_to(message, f"El monto de ${monto_compra} USD supera la franquicia máxima de $1000 USD permitida.")
                bot.send_message(usuario_id, "Para este envío necesitas tramitarlo por el Régimen General con un Despachante de Aduana.")
            
            #Fin del proceso, se reinicia el estado
            bot.send_message(usuario_id, "Escribí /start para hacer una nueva consulta.")
            actualizar_estado_excel(usuario_id, 'inicio')
            
        except ValueError:
            bot.reply_to(message, "Por favor, ingresá solo números (podés usar un punto para los decimales).")

#Ejecucion
if __name__ == '__main__':
    print("El bot de logística está en funcionamiento...")
    bot.infinity_polling()

